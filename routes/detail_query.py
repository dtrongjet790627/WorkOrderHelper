# -*- coding: utf-8 -*-
"""明细查询相关路由"""

from flask import Blueprint, request, jsonify, Response
from config.database import LINE_CONFIG
from models.acc_db import get_connection
from utils.line_identifier import identify_line
from utils.deployment import check_line_access
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

detail_query_bp = Blueprint('detail_query', __name__)


@detail_query_bp.route('/api/detail/pack_list', methods=['GET'])
def get_pack_list():
    """获取工单的包装批次列表"""
    try:
        wono = request.args.get('wono', '').strip()
        if not wono:
            return jsonify({'error': '请提供工单号'}), 400

        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        line_info = LINE_CONFIG[line_key]

        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询工单基本信息
        cursor.execute("""
            SELECT MAX(partno), MAX(line)
            FROM acc_wo_workorder_detail
            WHERE wono = :wono
        """, {'wono': wono})
        wo_row = cursor.fetchone()
        if not wo_row or not wo_row[0]:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到工单 {wono}'}), 404

        wo_line = wo_row[1]

        # 查询包装批次列表
        # pack_info.status: 0=打包中, 2=已封包
        cursor.execute("""
            SELECT DISTINCT
                pi.packid,
                pi.currquantity,
                pi.status,
                TO_CHAR(pi.lastupdatetime, 'YYYY-MM-DD HH24:MI:SS') AS lastupdate,
                (SELECT COUNT(*) FROM pack_history ph WHERE ph.packid = pi.packid AND ph.line = :wo_line) AS pack_count,
                (SELECT rs.SCHB_NUMBER FROM ACC_ERP_REPORT_SUCCESS rs
                 WHERE rs.PACKID = pi.packid AND rs.IS_SUCCESS = 1 AND ROWNUM = 1) AS schb_number
            FROM pack_history ph
            JOIN pack_info pi ON ph.packid = pi.packid
            WHERE EXISTS (
                SELECT 1 FROM acc_wo_workorder_detail awd
                WHERE awd.unitsn = ph.unitsn
                AND awd.wono = :wono
                AND awd.line = ph.line
            )
            AND ph.line = :wo_line
            ORDER BY pi.packid
        """, {'wono': wono, 'wo_line': wo_line})

        pack_list = []
        for row in cursor.fetchall():
            status_code = str(row[2]) if row[2] is not None else '0'
            if status_code == '2':
                status_text = '已封包'
            elif status_code == '0':
                status_text = '打包中'
            else:
                status_text = '未知'

            # 判断是否已报工
            if row[5]:
                status_text = '已报工'

            pack_list.append({
                'packid': row[0],
                'quantity': row[1] or row[4],  # 优先使用currquantity，没有则用pack_count
                'status': status_text,
                'status_code': status_code,
                'lastupdate': row[3],
                'schb_number': row[5]
            })

        cursor.close()
        conn.close()

        return jsonify({
            'wono': wono,
            'line': wo_line,
            'line_name': line_info['name'],
            'packs': pack_list,
            'total_packs': len(pack_list)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/pack_detail', methods=['GET'])
def get_pack_detail():
    """获取包装批次的详细信息（产品列表）"""
    try:
        packid = request.args.get('packid', '').strip()
        line = request.args.get('line', '').strip()

        if not packid:
            return jsonify({'error': '请提供批次号'}), 400

        # 根据packid格式或line参数识别数据库连接
        # 注意：数据库中line字段可能是"SMT Line2"格式，也可能是"SMT-2"格式
        if not line:
            # 尝试从packid识别
            packid_upper = packid.upper()
            if 'SMT-2' in packid_upper or 'MID-2' in packid_upper or 'SMT LINE2' in packid_upper or 'MID LINE2' in packid_upper:
                line_key = 'smt2'
            elif 'EPS' in packid_upper or 'IPA' in packid_upper:
                line_key = 'dpeps1'
            else:
                line_key = 'dpepp1'
        else:
            # 根据传入的line参数识别
            line_upper = line.upper()
            if 'SMT-2' in line_upper or 'MID-2' in line_upper or 'SMT LINE2' in line_upper or 'MID LINE2' in line_upper:
                line_key = 'smt2'
            elif 'EPS' in line_upper or 'IPA' in line_upper:
                line_key = 'dpeps1'
            else:
                line_key = 'dpepp1'

        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询批次基本信息（pack_info表没有maxquantity字段）
        cursor.execute("""
            SELECT
                packid,
                currquantity,
                packsize,
                status,
                TO_CHAR(lastupdatetime, 'YYYY-MM-DD HH24:MI:SS') AS lastupdate,
                prodtype
            FROM pack_info
            WHERE packid = :packid
        """, {'packid': packid})
        pack_row = cursor.fetchone()

        if not pack_row:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到批次 {packid}'}), 404

        status_code = str(pack_row[3]) if pack_row[3] is not None else '0'
        if status_code == '2':
            status_text = '已封包'
        elif status_code == '0':
            status_text = '打包中'
        else:
            status_text = '未知'

        # 查询报工单号
        cursor.execute("""
            SELECT SCHB_NUMBER FROM ACC_ERP_REPORT_SUCCESS
            WHERE PACKID = :packid AND IS_SUCCESS = 1 AND ROWNUM = 1
        """, {'packid': packid})
        schb_row = cursor.fetchone()
        schb_number = schb_row[0] if schb_row else None

        if schb_number:
            status_text = '已报工'

        pack_info = {
            'packid': pack_row[0],
            'currquantity': pack_row[1],
            'packsize': pack_row[2],
            'status': status_text,
            'status_code': status_code,
            'lastupdate': pack_row[4],
            'prodtype': pack_row[5],
            'schb_number': schb_number
        }

        # 查询批次内的产品列表（限制300条）
        # 优化：移除LEFT JOIN，直接从pack_history获取packdate和stn，状态从前端可选查询
        query_params = {'packid': packid}
        line_filter = ""
        if line:
            line_filter = "AND ph.line = :line"
            query_params['line'] = line

        cursor.execute(f"""
            SELECT * FROM (
                SELECT
                    ph.unitsn,
                    ph.line,
                    TO_CHAR(ph.packdate, 'YYYY-MM-DD HH24:MI:SS') AS packdate,
                    ph.stn
                FROM pack_history ph
                WHERE ph.packid = :packid {line_filter}
                ORDER BY ph.packdate, ph.unitsn
            ) WHERE ROWNUM <= 300
        """, query_params)

        products = []
        for idx, row in enumerate(cursor.fetchall(), 1):
            # row: unitsn, line, packdate, stn
            products.append({
                'seq': idx,
                'unitsn': row[0],
                'line': row[1],
                'packdate': row[2],
                'stn': row[3]
            })

        cursor.close()
        conn.close()

        return jsonify({
            'pack_info': pack_info,
            'products': products,
            'total_count': len(products)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/unit_trace', methods=['GET'])
def get_unit_trace():
    """获取单品过站追溯信息

    重要：必须严格按传入的line参数查询，确保只查询当前工单产线的数据
    同一个SN可能在多条产线都有记录，必须用line过滤
    """
    try:
        unitsn = request.args.get('unitsn', '').strip()
        line = request.args.get('line', '').strip()

        if not unitsn:
            return jsonify({'error': '请提供产品序列号'}), 400

        # 根据传入的line识别数据库连接
        # 注意：数据库中line字段可能是"SMT Line2"格式，也可能是"SMT-2"格式
        if line:
            line_upper = line.upper()
            if 'SMT-2' in line_upper or 'MID-2' in line_upper or 'SMT LINE2' in line_upper or 'MID LINE2' in line_upper:
                line_key = 'smt2'
            elif 'EPS' in line_upper or 'IPA' in line_upper:
                line_key = 'dpeps1'
            else:
                line_key = 'dpepp1'
        else:
            # 默认电控一线
            line_key = 'dpepp1'

        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 必须使用传入的line参数进行过滤，确保只查询当前工单产线的数据
        query_line = line if line else None

        # 先查询acc_unitstatus（当前状态表）
        if query_line:
            cursor.execute("""
                SELECT
                    us.unitsn,
                    us.partno,
                    us.line,
                    us.status,
                    us.op
                FROM acc_unitstatus us
                WHERE us.unitsn = :unitsn AND us.line = :line
            """, {'unitsn': unitsn, 'line': query_line})
        else:
            cursor.execute("""
                SELECT
                    us.unitsn,
                    us.partno,
                    us.line,
                    us.status,
                    us.op
                FROM acc_unitstatus us
                WHERE us.unitsn = :unitsn
            """, {'unitsn': unitsn})
        unit_row = cursor.fetchone()

        # 如果acc_unitstatus没有记录，尝试从acc_unithistory获取基本信息
        # （产品可能还在生产中，尚未写入unitstatus）
        if not unit_row and query_line:
            cursor.execute("""
                SELECT DISTINCT
                    uh.unitsn,
                    uh.partno,
                    uh.line,
                    NULL AS status,
                    MAX(uh.op) AS op
                FROM acc_unithistory uh
                WHERE uh.unitsn = :unitsn AND uh.line = :line
                GROUP BY uh.unitsn, uh.partno, uh.line
            """, {'unitsn': unitsn, 'line': query_line})
            unit_row = cursor.fetchone()

        if not unit_row:
            cursor.close()
            conn.close()
            if query_line:
                return jsonify({'error': f'产品 {unitsn} 在产线 {query_line} 无记录'}), 404
            else:
                return jsonify({'error': f'未找到产品 {unitsn}'}), 404

        unit_status = unit_row[3]
        # 转换为整数进行比较（Oracle可能返回Decimal类型）
        try:
            status_val = int(unit_status) if unit_status is not None else None
        except (ValueError, TypeError):
            status_val = None

        if status_val == 2:
            status_text = 'FG'
        elif status_val == 1:
            status_text = 'WIP'
        else:
            status_text = '-'

        # 使用查询结果中的line（应该与传入的line一致）
        unit_line = unit_row[2]

        # 查询所属批次
        cursor.execute("""
            SELECT packid FROM pack_history
            WHERE unitsn = :unitsn AND line = :line AND ROWNUM = 1
        """, {'unitsn': unitsn, 'line': unit_line})
        pack_row = cursor.fetchone()
        packid = pack_row[0] if pack_row else None

        # 查询工单号（从acc_wo_workorder_detail表）
        cursor.execute("""
            SELECT wono FROM acc_wo_workorder_detail
            WHERE unitsn = :unitsn AND line = :line AND ROWNUM = 1
        """, {'unitsn': unitsn, 'line': unit_line})
        wono_row = cursor.fetchone()
        wono = wono_row[0] if wono_row else None

        unit_info = {
            'unitsn': unit_row[0],
            'partno': unit_row[1],
            'line': unit_line,
            'status': status_text,
            'status_code': status_val,
            'current_op': unit_row[4],
            'packid': packid,
            'wono': wono
        }

        # 查询过站履历
        cursor.execute("""
            SELECT
                op,
                TO_CHAR(STARTDT, 'YYYY-MM-DD HH24:MI:SS') AS start_time,
                TO_CHAR(ENDDT, 'YYYY-MM-DD HH24:MI:SS') AS end_time,
                result,
                partno
            FROM acc_unithistory
            WHERE unitsn = :unitsn AND line = :line
            ORDER BY STARTDT
        """, {'unitsn': unitsn, 'line': unit_line})

        history = []
        for idx, row in enumerate(cursor.fetchall(), 1):
            result_code = row[3]
            # 转换为整数进行比较（Oracle可能返回Decimal类型）
            try:
                result_val = int(result_code) if result_code is not None else None
            except (ValueError, TypeError):
                result_val = None

            # result: 1=OK, 0=NOK, 2=N/A
            if result_val == 1:
                result_text = 'OK'
            elif result_val == 0:
                result_text = 'NOK'
            elif result_val == 2:
                result_text = 'N/A'
            else:
                result_text = '-'

            history.append({
                'seq': idx,
                'op': row[0],
                'start_time': row[1],
                'end_time': row[2],
                'result': result_text,
                'result_code': result_val,
                'partno': row[4]
            })

        cursor.close()
        conn.close()

        return jsonify({
            'unit_info': unit_info,
            'history': history,
            'total_stations': len(history)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/finished_products', methods=['GET'])
def get_finished_products():
    """获取工单的完工产品明细（FG状态）

    返回数据结构：
    - summary: 统计卡片（完工总数、已打包、未打包、型号、末站工位）
    - products: 产品列表（序号、产品SN、完工时间、包装状态、返工次数）

    性能优化说明：
    - 使用两步查询避免复杂的多表JOIN
    - 先查询完工产品SN列表，再批量查询打包信息
    - 使用IN子句替代NOT EXISTS提升性能
    """
    try:
        wono = request.args.get('wono', '').strip()
        if not wono:
            return jsonify({'error': '请提供工单号'}), 400

        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询工单基本信息
        cursor.execute("""
            SELECT MAX(partno), MAX(line)
            FROM acc_wo_workorder_detail
            WHERE wono = :wono
        """, {'wono': wono})
        wo_row = cursor.fetchone()
        if not wo_row or not wo_row[0]:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到工单 {wono}'}), 404

        partno = wo_row[0]
        wo_line = wo_row[1]

        # 步骤1：查询工单内的完工产品SN列表（只从acc_wo_workorder_detail查询）
        # 使用status=2筛选完工产品，避免JOIN acc_unitstatus
        cursor.execute("""
            SELECT unitsn
            FROM acc_wo_workorder_detail
            WHERE wono = :wono AND status = 2 AND line = :wo_line
            ORDER BY unitsn
        """, {'wono': wono, 'wo_line': wo_line})

        finished_sns = [row[0] for row in cursor.fetchall()]
        total_count = len(finished_sns)

        if total_count == 0:
            # 没有完工产品，直接返回空结果
            # 查询在制数量
            cursor.execute("""
                SELECT COUNT(*)
                FROM acc_wo_workorder_detail
                WHERE wono = :wono AND status = 1 AND line = :wo_line
            """, {'wono': wono, 'wo_line': wo_line})
            wip_row = cursor.fetchone()
            wip_count = wip_row[0] if wip_row else 0

            cursor.close()
            conn.close()
            return jsonify({
                'summary': {
                    'total_count': 0,
                    'packed_count': 0,
                    'unpacked_count': 0,
                    'wip_count': wip_count,
                    'partno': partno,
                    'last_station': '-'
                },
                'products': []
            })

        # 步骤2：批量查询这些SN的打包信息
        # 使用临时表或IN子句（Oracle对IN有限制，分批处理）
        pack_info = {}  # {unitsn: packid}

        # Oracle的IN子句最多支持1000个元素，分批查询
        batch_size = 900
        for i in range(0, len(finished_sns), batch_size):
            batch = finished_sns[i:i+batch_size]
            # 创建绑定变量
            bind_names = [f':sn{j}' for j in range(len(batch))]
            bind_values = {f'sn{j}': sn for j, sn in enumerate(batch)}
            bind_values['wo_line'] = wo_line

            cursor.execute(f"""
                SELECT unitsn, packid
                FROM pack_history
                WHERE unitsn IN ({','.join(bind_names)}) AND line = :wo_line
            """, bind_values)

            for row in cursor.fetchall():
                pack_info[row[0]] = row[1]

        # 步骤3：批量查询完工时间（从acc_unitstatus表的DT字段获取）
        finish_time_info = {}  # {unitsn: dt}
        for i in range(0, len(finished_sns), batch_size):
            batch = finished_sns[i:i+batch_size]
            bind_names = [f':sn{j}' for j in range(len(batch))]
            bind_values = {f'sn{j}': sn for j, sn in enumerate(batch)}
            bind_values['wo_line'] = wo_line

            cursor.execute(f"""
                SELECT unitsn, TO_CHAR(dt, 'YYYY-MM-DD HH24:MI:SS')
                FROM acc_unitstatus
                WHERE unitsn IN ({','.join(bind_names)}) AND line = :wo_line
            """, bind_values)

            for row in cursor.fetchall():
                finish_time_info[row[0]] = row[1]

        # 步骤4：查询末站工位（从第一个完工产品获取）
        cursor.execute("""
            SELECT op FROM acc_unitstatus
            WHERE unitsn = :unitsn AND line = :wo_line
        """, {'unitsn': finished_sns[0], 'wo_line': wo_line})
        op_row = cursor.fetchone()
        last_station = op_row[0] if op_row else '-'

        # 构建产品列表
        products = []
        packed_count = 0
        for unitsn in finished_sns:
            packid = pack_info.get(unitsn, '-')
            if packid and packid != '-':
                packed_count += 1

            # 完工时间从acc_unitstatus的DT字段获取
            finish_time = finish_time_info.get(unitsn, '-')

            products.append({
                'unitsn': unitsn,
                'finish_time': finish_time or '-',
                'packid': packid or '-'
            })

        unpacked_count = total_count - packed_count

        # 排序：未打包的排在前面，然后按完工时间降序（最新的在前面）
        def sort_key(p):
            # 未打包(packid为'-'或空)排在前面
            is_packed = 1 if (p['packid'] and p['packid'] != '-') else 0
            # 完工时间降序（空值排最后，用0000开头；有值的取反排序）
            ft = p['finish_time'] if p['finish_time'] != '-' else '0000-00-00'
            return (is_packed, ft)

        products.sort(key=sort_key, reverse=False)
        # 分组排序：未打包在前按时间降序，已打包在后按时间降序
        unpacked = [p for p in products if not p['packid'] or p['packid'] == '-']
        packed = [p for p in products if p['packid'] and p['packid'] != '-']
        unpacked.sort(key=lambda p: p['finish_time'] if p['finish_time'] != '-' else '0000-00-00', reverse=True)
        packed.sort(key=lambda p: p['finish_time'] if p['finish_time'] != '-' else '0000-00-00', reverse=True)
        products = unpacked + packed

        # 重新编号
        for idx, p in enumerate(products, 1):
            p['seq'] = idx

        # 查询在制数量
        cursor.execute("""
            SELECT COUNT(*)
            FROM acc_wo_workorder_detail
            WHERE wono = :wono AND status = 1 AND line = :wo_line
        """, {'wono': wono, 'wo_line': wo_line})
        wip_row = cursor.fetchone()
        wip_count = wip_row[0] if wip_row else 0

        cursor.close()
        conn.close()

        return jsonify({
            'summary': {
                'total_count': total_count,
                'packed_count': packed_count,
                'unpacked_count': unpacked_count,
                'wip_count': wip_count,
                'partno': partno,
                'last_station': last_station or '-'
            },
            'products': products
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/wip_products', methods=['GET'])
def get_wip_products():
    """获取工单的在制产品明细（WIP状态）

    返回数据结构：
    - summary: 统计卡片（在制总数）
    - products: 产品列表（序号、产品SN、当前工位、进站时间、滞留时长）
    """
    try:
        wono = request.args.get('wono', '').strip()
        if not wono:
            return jsonify({'error': '请提供工单号'}), 400

        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询工单产线（优化：用ROWNUM=1代替MAX，避免全表扫描）
        cursor.execute("""
            SELECT line FROM acc_wo_workorder_detail
            WHERE wono = :wono AND ROWNUM = 1
        """, {'wono': wono})
        wo_row = cursor.fetchone()
        if not wo_row or not wo_row[0]:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到工单 {wono}'}), 404

        wo_line = wo_row[0]

        # 第一步：获取在制品SN列表
        cursor.execute("""
            SELECT unitsn
            FROM acc_wo_workorder_detail
            WHERE wono = :wono AND status = 1 AND line = :wo_line
        """, {'wono': wono, 'wo_line': wo_line})
        sn_list = [row[0] for row in cursor.fetchall()]

        from datetime import datetime
        now = datetime.now()
        products = []

        # 第二步：对每个SN查询最新历史记录（用ROWNUM=1，兼容旧版Oracle）
        for idx, sn in enumerate(sn_list, 1):
            cursor.execute("""
                SELECT * FROM (
                    SELECT op, STARTDT, ENDDT, result
                    FROM acc_unithistory
                    WHERE unitsn = :sn AND line = :wo_line
                    ORDER BY STARTDT DESC
                ) WHERE ROWNUM = 1
            """, {'sn': sn, 'wo_line': wo_line})
            hist_row = cursor.fetchone()

            current_op = '-'
            enter_time_str = '-'
            end_time_str = '-'
            result_str = '-'
            duration_str = '-'

            if hist_row:
                current_op = hist_row[0] or '-'
                enter_dt = hist_row[1]
                end_dt = hist_row[2]
                result_val = hist_row[3]

                # 加工结果（Oracle返回Decimal类型，需转为int匹配字典键）
                if result_val is not None:
                    try:
                        result_int = int(result_val)
                        result_str = {0: 'NOK', 1: 'OK', 2: 'N/A'}.get(result_int, str(result_val))
                    except (ValueError, TypeError):
                        result_str = str(result_val)

                # 进站时间
                if enter_dt:
                    try:
                        enter_time_str = enter_dt.strftime('%Y-%m-%d %H:%M:%S')
                        diff = now - enter_dt
                        total_minutes = int(diff.total_seconds() / 60)
                        if total_minutes < 60:
                            duration_str = f'{total_minutes}分钟'
                        elif total_minutes < 1440:
                            hours = total_minutes // 60
                            mins = total_minutes % 60
                            duration_str = f'{hours}小时{mins}分钟'
                        else:
                            days = total_minutes // 1440
                            hours = (total_minutes % 1440) // 60
                            duration_str = f'{days}天{hours}小时'
                    except:
                        pass

                # 出站时间
                if end_dt:
                    try:
                        end_time_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass

            products.append({
                'seq': idx,
                'unitsn': sn,
                'current_op': current_op,
                'enter_time': enter_time_str,
                'end_time': end_time_str,
                'result': result_str,
                'duration': duration_str
            })

        total_count = len(products)

        cursor.close()
        conn.close()

        return jsonify({
            'summary': {
                'total_count': total_count
            },
            'products': products
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def create_excel_workbook(headers, data, column_widths=None):
    """创建带样式的Excel工作簿

    Args:
        headers: 表头列表
        data: 数据行列表（每行是一个列表）
        column_widths: 列宽列表（可选）

    Returns:
        Workbook对象
    """
    wb = Workbook()
    ws = wb.active

    # 定义样式
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    even_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    data_alignment = Alignment(horizontal='left', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, row_data in enumerate(data, 2):
        row_fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = data_alignment
            cell.border = thin_border

    # 设置列宽
    if column_widths:
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        # 自动调整列宽（基于内容）
        for col_idx in range(1, len(headers) + 1):
            max_length = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            # 设置列宽，增加一些边距
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    return wb


@detail_query_bp.route('/api/detail/export_packs', methods=['GET'])
def export_packs():
    """导出工单的所有包装批次（汇总信息）"""
    try:
        wono = request.args.get('wono', '').strip()
        if not wono:
            return jsonify({'error': '请提供工单号'}), 400

        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询工单的line和型号
        cursor.execute("""
            SELECT MAX(line), MAX(partno) FROM acc_wo_workorder_detail WHERE wono = :wono
        """, {'wono': wono})
        wo_row = cursor.fetchone()
        if not wo_row or not wo_row[0]:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到工单 {wono}'}), 404

        wo_line = wo_row[0]
        partno = wo_row[1] or '-'

        # 查询包装汇总信息（不再遍历每个产品）
        cursor.execute("""
            SELECT DISTINCT
                pi.packid,
                pi.PRODTYPE,
                pi.currquantity,
                pi.status,
                TO_CHAR(pi.lastupdatetime, 'YYYY-MM-DD HH24:MI:SS') AS lastupdate,
                (SELECT rs.SCHB_NUMBER FROM ACC_ERP_REPORT_SUCCESS rs
                 WHERE rs.PACKID = pi.packid AND rs.IS_SUCCESS = 1 AND ROWNUM = 1) AS schb_number
            FROM pack_history ph
            JOIN pack_info pi ON ph.packid = pi.packid
            WHERE EXISTS (
                SELECT 1 FROM acc_wo_workorder_detail awd
                WHERE awd.unitsn = ph.unitsn
                AND awd.wono = :wono
                AND awd.line = ph.line
            )
            AND ph.line = :wo_line
            ORDER BY pi.packid
        """, {'wono': wono, 'wo_line': wo_line})

        # 收集数据并统计
        rows_data = cursor.fetchall()
        total_packs = len(rows_data)
        total_products = sum((row[2] or 0) for row in rows_data)

        cursor.close()
        conn.close()

        # 创建Excel（带统计信息头部）
        wb = Workbook()
        ws = wb.active

        # 定义样式
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        info_font = Font(bold=True)
        info_alignment = Alignment(horizontal='left', vertical='center')
        odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        even_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        data_alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # 写入统计信息（前4行）
        info_rows = [
            ('工单号：', wono),
            ('型号：', partno),
            ('包装数：', f'{total_packs} 个'),
            ('产品总数：', f'{total_products} 个'),
        ]
        for row_idx, (label, value) in enumerate(info_rows, 1):
            ws.cell(row=row_idx, column=1, value=label).font = info_font
            ws.cell(row=row_idx, column=1).alignment = info_alignment
            ws.cell(row=row_idx, column=2, value=value)

        # 空一行后写入表头（第6行）
        headers = ['序号', '批次号', '型号', '数量', '状态', '更新时间', '报工单号']
        header_row = 6
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for seq, row in enumerate(rows_data, 1):
            row_idx = header_row + seq
            # row: packid, prodtype, currquantity, status, lastupdate, schb_number
            status_code = str(row[3]) if row[3] is not None else '0'
            if row[5]:
                status_text = '已报工'
            elif status_code == '2':
                status_text = '已封包'
            elif status_code == '0':
                status_text = '打包中'
            else:
                status_text = '未知'

            row_data = [seq, row[0] or '-', row[1] or '-', row[2] or 0, status_text, row[4] or '-', row[5] or '-']
            row_fill = even_row_fill if seq % 2 == 0 else odd_row_fill

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.alignment = data_alignment
                cell.border = thin_border

        # 设置列宽
        column_widths = [8, 25, 20, 10, 10, 22, 20]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=pack_list_{wono}.xlsx'}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/export_pack_detail', methods=['GET'])
def export_pack_detail():
    """导出单个批次的产品明细"""
    try:
        packid = request.args.get('packid', '').strip()
        line = request.args.get('line', '').strip()
        wono = request.args.get('wono', '').strip()  # 工单号参数

        if not packid:
            return jsonify({'error': '请提供批次号'}), 400

        # 根据packid格式或line参数识别数据库连接
        if not line:
            packid_upper = packid.upper()
            if 'SMT-2' in packid_upper or 'MID-2' in packid_upper or 'SMT LINE2' in packid_upper or 'MID LINE2' in packid_upper:
                line_key = 'smt2'
            elif 'EPS' in packid_upper or 'IPA' in packid_upper:
                line_key = 'dpeps1'
            else:
                line_key = 'dpepp1'
        else:
            line_upper = line.upper()
            if 'SMT-2' in line_upper or 'MID-2' in line_upper or 'SMT LINE2' in line_upper or 'MID LINE2' in line_upper:
                line_key = 'smt2'
            elif 'EPS' in line_upper or 'IPA' in line_upper:
                line_key = 'dpeps1'
            else:
                line_key = 'dpepp1'

        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 如果没有传工单号，从数据库查询
        if not wono:
            cursor.execute("""
                SELECT awd.wono FROM pack_history ph
                JOIN acc_wo_workorder_detail awd ON ph.unitsn = awd.unitsn AND ph.line = awd.line
                WHERE ph.packid = :packid AND ROWNUM = 1
            """, {'packid': packid})
            wono_row = cursor.fetchone()
            wono = wono_row[0] if wono_row else '-'

        # 查询批次内的所有产品（partno从acc_unitstatus获取）
        cursor.execute("""
            SELECT
                ph.unitsn,
                us.partno,
                ph.line,
                TO_CHAR(ph.packdate, 'YYYY-MM-DD HH24:MI:SS') AS packdate
            FROM pack_history ph
            LEFT JOIN acc_unitstatus us ON ph.unitsn = us.unitsn AND ph.line = us.line
            WHERE ph.packid = :packid
            ORDER BY ph.packdate
        """, {'packid': packid})

        # 准备Excel数据（不含状态字段）
        # 列顺序：序号、产品序列号、型号、产线、工单号、批次号、打包时间
        headers = ['序号', '产品序列号', '型号', '产线', '工单号', '批次号', '打包时间']
        data = []

        for idx, row in enumerate(cursor.fetchall(), 1):
            # row: unitsn, partno, line, packdate
            data.append([idx, row[0] or '-', row[1] or '-', row[2] or '-', wono, packid, row[3] or '-'])

        cursor.close()
        conn.close()

        # 创建Excel
        wb = create_excel_workbook(headers, data, column_widths=[8, 30, 20, 15, 20, 25, 22])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=pack_detail_{packid}.xlsx'}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/export_finished', methods=['GET'])
def export_finished_products():
    """导出工单的完工产品明细

    性能优化说明：
    - 使用两步查询避免复杂的多表JOIN
    - 先查询完工产品SN列表，再批量查询打包信息
    """
    try:
        wono = request.args.get('wono', '').strip()
        if not wono:
            return jsonify({'error': '请提供工单号'}), 400

        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询工单的line
        cursor.execute("""
            SELECT MAX(line), MAX(partno) FROM acc_wo_workorder_detail WHERE wono = :wono
        """, {'wono': wono})
        wo_row = cursor.fetchone()
        if not wo_row or not wo_row[0]:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到工单 {wono}'}), 404

        wo_line = wo_row[0]
        partno = wo_row[1]

        # 步骤1：查询完工产品SN列表
        cursor.execute("""
            SELECT unitsn
            FROM acc_wo_workorder_detail
            WHERE wono = :wono AND status = 2 AND line = :wo_line
            ORDER BY unitsn
        """, {'wono': wono, 'wo_line': wo_line})

        finished_sns = [row[0] for row in cursor.fetchall()]

        # 步骤2：批量查询打包信息
        pack_info = {}  # {unitsn: packid}
        batch_size = 900
        for i in range(0, len(finished_sns), batch_size):
            batch = finished_sns[i:i+batch_size]
            bind_names = [f':sn{j}' for j in range(len(batch))]
            bind_values = {f'sn{j}': sn for j, sn in enumerate(batch)}
            bind_values['wo_line'] = wo_line

            cursor.execute(f"""
                SELECT unitsn, packid
                FROM pack_history
                WHERE unitsn IN ({','.join(bind_names)}) AND line = :wo_line
            """, bind_values)

            for row in cursor.fetchall():
                pack_info[row[0]] = row[1]

        # 步骤3：批量查询完工时间（从acc_unitstatus表的DT字段获取）
        finish_time_info = {}  # {unitsn: dt}
        for i in range(0, len(finished_sns), batch_size):
            batch = finished_sns[i:i+batch_size]
            bind_names = [f':sn{j}' for j in range(len(batch))]
            bind_values = {f'sn{j}': sn for j, sn in enumerate(batch)}
            bind_values['wo_line'] = wo_line

            cursor.execute(f"""
                SELECT unitsn, TO_CHAR(dt, 'YYYY-MM-DD HH24:MI:SS')
                FROM acc_unitstatus
                WHERE unitsn IN ({','.join(bind_names)}) AND line = :wo_line
            """, bind_values)

            for row in cursor.fetchall():
                finish_time_info[row[0]] = row[1]

        cursor.close()
        conn.close()

        # 构建产品列表并排序
        products = []
        for unitsn in finished_sns:
            packid = pack_info.get(unitsn, '-')
            finish_time = finish_time_info.get(unitsn, '-')
            products.append({
                'unitsn': unitsn,
                'finish_time': finish_time or '-',
                'packid': packid or '-'
            })

        # 排序：未打包的排在前面，然后按完工时间降序（最新的在前面）
        unpacked = [p for p in products if not p['packid'] or p['packid'] == '-']
        packed = [p for p in products if p['packid'] and p['packid'] != '-']
        unpacked.sort(key=lambda p: p['finish_time'] if p['finish_time'] != '-' else '0000-00-00', reverse=True)
        packed.sort(key=lambda p: p['finish_time'] if p['finish_time'] != '-' else '0000-00-00', reverse=True)
        products = unpacked + packed

        # 准备Excel数据
        # 列顺序：序号、产品序列号、型号、工单号、完工时间、批次
        headers = ['序号', '产品序列号', '型号', '工单号', '完工时间', '批次']
        data = []

        for idx, p in enumerate(products, 1):
            data.append([idx, p['unitsn'], partno or '-', wono, p['finish_time'], p['packid']])

        # 创建Excel
        wb = create_excel_workbook(headers, data, column_widths=[8, 30, 20, 20, 22, 25])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=finished_products_{wono}.xlsx'}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/detail/export_wip', methods=['GET'])
def export_wip_products():
    """导出工单的在制产品明细"""
    try:
        wono = request.args.get('wono', '').strip()
        if not wono:
            return jsonify({'error': '请提供工单号'}), 400

        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        conn = get_connection(line_key)
        cursor = conn.cursor()

        # 查询工单的line
        cursor.execute("""
            SELECT MAX(line), MAX(partno) FROM acc_wo_workorder_detail WHERE wono = :wono
        """, {'wono': wono})
        wo_row = cursor.fetchone()
        if not wo_row or not wo_row[0]:
            cursor.close()
            conn.close()
            return jsonify({'error': f'未找到工单 {wono}'}), 404

        wo_line = wo_row[0]
        partno = wo_row[1]

        # 查询在制产品（以acc_wo_workorder_detail.status=1为准，与查询接口一致）
        cursor.execute("""
            SELECT
                awd.unitsn,
                us.op AS current_op,
                MAX(uh.startdt) AS enter_dt
            FROM acc_wo_workorder_detail awd
            LEFT JOIN acc_unitstatus us ON us.unitsn = awd.unitsn AND us.line = awd.line
            LEFT JOIN acc_unithistory uh ON uh.unitsn = awd.unitsn AND uh.line = awd.line AND uh.op = us.op
            WHERE awd.wono = :wono AND awd.status = 1 AND awd.line = :wo_line
            GROUP BY awd.unitsn, us.op
            ORDER BY awd.unitsn
        """, {'wono': wono, 'wo_line': wo_line})

        from datetime import datetime
        now = datetime.now()

        # 准备Excel数据
        # 列顺序：序号、产品序列号、型号、工单号、当前工位、进站时间、滞留时长
        headers = ['序号', '产品序列号', '型号', '工单号', '当前工位', '进站时间', '滞留时长']
        data = []

        for idx, row in enumerate(cursor.fetchall(), 1):
            # row: unitsn, current_op, enter_dt
            enter_dt = row[2]
            enter_time_str = '-'
            duration_str = '-'

            if enter_dt:
                try:
                    enter_time_str = enter_dt.strftime('%Y-%m-%d %H:%M:%S')
                    diff = now - enter_dt
                    total_minutes = int(diff.total_seconds() / 60)
                    if total_minutes < 60:
                        duration_str = f'{total_minutes}分钟'
                    elif total_minutes < 1440:
                        hours = total_minutes // 60
                        mins = total_minutes % 60
                        duration_str = f'{hours}小时{mins}分钟'
                    else:
                        days = total_minutes // 1440
                        hours = (total_minutes % 1440) // 60
                        duration_str = f'{days}天{hours}小时'
                except:
                    pass

            data.append([idx, row[0] or '-', partno or '-', wono, row[1] or '-', enter_time_str, duration_str])

        cursor.close()
        conn.close()

        # 创建Excel
        wb = create_excel_workbook(headers, data, column_widths=[8, 30, 20, 20, 15, 22, 15])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=wip_products_{wono}.xlsx'}
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@detail_query_bp.route('/api/debug/check_indexes', methods=['GET'])
def check_indexes():
    """检查数据库表的索引情况（调试用）"""
    try:
        wono = request.args.get('wono', 'MID-226011401').strip()
        line_key = identify_line(wono)
        allowed, err = check_line_access(line_key)
        if not allowed:
            return err
        conn = get_connection(line_key)
        cursor = conn.cursor()

        result = {}

        # 检查 acc_unithistory 表索引
        cursor.execute("""
            SELECT index_name, column_name, column_position
            FROM all_ind_columns
            WHERE table_name = 'ACC_UNITHISTORY'
            ORDER BY index_name, column_position
        """)
        result['acc_unithistory_indexes'] = [
            {'index': row[0], 'column': row[1], 'pos': row[2]}
            for row in cursor.fetchall()
        ]

        # 检查 acc_wo_workorder_detail 表索引
        cursor.execute("""
            SELECT index_name, column_name, column_position
            FROM all_ind_columns
            WHERE table_name = 'ACC_WO_WORKORDER_DETAIL'
            ORDER BY index_name, column_position
        """)
        result['acc_wo_workorder_detail_indexes'] = [
            {'index': row[0], 'column': row[1], 'pos': row[2]}
            for row in cursor.fetchall()
        ]

        cursor.close()
        conn.close()

        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500
